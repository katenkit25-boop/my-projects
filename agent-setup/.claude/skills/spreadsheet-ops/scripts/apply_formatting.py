"""Format Excel files (numeric format, column width, alignment).
Configure formatting rules in CONFIG."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CONFIG = {
    "input_file": "input.xlsx",
    "output_file": "output_with_formatting.xlsx",
    "sheet_name": "Sheet1",
    "column_formats": {
        "C": {"number_format": "#,##0.00"},
        "D": {"number_format": "0.00%"},
    },
    "column_widths": {
        "A": 20,
        "B": 18,
    },
    "header_style": {
        "font_bold": True,
        "fill_color": "FFEEEEEE",
        "alignment": "center",
    },
}


def apply_header_style(sheet) -> None:
    if not CONFIG.get("header_style"):
        return
    style = CONFIG["header_style"]
    fill = PatternFill("solid", fgColor=style["fill_color"])
    font = Font(bold=style.get("font_bold", False))
    alignment = Alignment(horizontal=style.get("alignment", "left"))
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def main() -> None:
    workbook = load_workbook(CONFIG["input_file"])
    sheet = workbook[CONFIG["sheet_name"]]

    for col, fmt in CONFIG["column_formats"].items():
        for cell in sheet[col]:
            cell.number_format = fmt["number_format"]

    for col, width in CONFIG["column_widths"].items():
        sheet.column_dimensions[col].width = width

    apply_header_style(sheet)

    output_path = Path(CONFIG["output_file"])
    workbook.save(output_path)
    print(f"Format writing completed: {output_path}")


if __name__ == "__main__":
    main()
