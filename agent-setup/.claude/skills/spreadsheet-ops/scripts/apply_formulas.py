"""Batch write formulas to Excel files.
Configure input files, target ranges, and formula templates in CONFIG."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


CONFIG = {
    "input_file": "input.xlsx",
    "output_file": "output_with_formulas.xlsx",
    "sheet_name": "Sheet1",
    "start_row": 2,
    "target_column": "D",
    "formula_template": "=B{row}*C{row}",
}


def main() -> None:
    workbook = load_workbook(CONFIG["input_file"])
    sheet = workbook[CONFIG["sheet_name"]]

    max_row = sheet.max_row
    for row in range(CONFIG["start_row"], max_row + 1):
        cell = f"{CONFIG['target_column']}{row}"
        sheet[cell].value = CONFIG["formula_template"].format(row=row)

    output_path = Path(CONFIG["output_file"])
    workbook.save(output_path)
    print(f"Formula writing completed: {output_path}")


if __name__ == "__main__":
    main()
